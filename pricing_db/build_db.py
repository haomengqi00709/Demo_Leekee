#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
解析李记包装全部价目 Excel -> SQLite (pricing.db)，并抽取内嵌图片映射到产品编码。
用法:  python3 build_db.py
"""
import os, re, json, sqlite3, glob
import openpyxl
import xlrd
import extract_images as imgmod

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   # bottle_pricing/
DBDIR = os.path.join(ROOT, 'pricing_db')
DB = os.path.join(DBDIR, 'pricing.db')
IMGDIR = os.path.join(DBDIR, 'images')

# 真实供应商（不含“李记”——李记是本公司；“李记型号”指本公司内部编码，不是供应商型号）
SUPPLIERS = ['丹之鸿','伟诚','瑞昶','派顿','雅利宝','云邦','正匠','日尚','恒邦','神龙',
             '万荣','威达','美顺','欧之美','新琪光','星日','天一','正兴','美妆优品','弘敏']

PROC_KW = ['实色','电镀','喷漆','uv','UV','水转印','透明','镀','烫','哑','漆','手感','触感',
           '幻彩','木纹','蒙砂','磨砂','光油','丝印','印刷','pp','PP','abs','ABS','银','金','色']

CODE_RE = re.compile(r'[A-Z]{1,3}\d{3,}[A-Z]?')     # CC20596 / CA10502 / H574 / SH2002 ...

# --------------------------------------------------------------------------
def norm(h):
    if h is None: return ''
    return re.sub(r'[\s/]+', '', str(h))

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    m = re.search(r'-?\d+\.?\d*', str(v).replace(',', ''))
    return float(m.group()) if m else None

def rawjson(cells):
    """序列化整行，但去掉尾部空单元格（避免 16000 列的表把 raw_json 撑到几十 KB）"""
    a = [('' if c is None else str(c)) for c in cells]
    while a and a[-1] in ('', 'None'):
        a.pop()
    return json.dumps(a, ensure_ascii=False)

def find_year(*texts):
    for t in texts:
        if not t: continue
        m = re.search(r'(20\d{2})', str(t))
        if m: return int(m.group(1))
    return None

def split_codes(raw):
    if not raw: return []
    return list(dict.fromkeys(CODE_RE.findall(str(raw).upper())))

NOTE_RE = re.compile(r'[：:，。]|说明|备注|注意|价格含|另加|以实际|谢谢|下单|如有|不含|提供')
def is_codelike(v):
    """判断某单元格是否像“产品编号”，而不是一段备注/说明文字"""
    if v is None: return False
    s = str(v).strip()
    if not s or len(s) > 30 or NOTE_RE.search(s): return False
    if CODE_RE.search(s.upper()): return True          # 含真实编码 CC/CA/H...
    if len(s) <= 12 and re.search(r'\d', s): return True  # 短且带数字: 001-02H / 18牙泵头 / H611
    return False

def is_num_only(v):
    return v is not None and re.fullmatch(r'[\d.\s]+', str(v)) is not None

def infer_type(prefix, code, name='', component='', process=''):
    ctx = f"{name}{component}{process}"
    if prefix == 'H':  return 'series'
    if prefix == 'CA': return 'cap'
    if prefix == 'CB': return 'bottle'
    if prefix == 'LB': return 'supplier_model'
    if '滴管' in ctx or '滴泵' in ctx: return 'dropper'
    if '泵' in ctx:  return 'pump'
    if '外罩' in ctx or '罩' in ctx: return 'cover'
    if '盖' in ctx:  return 'cap'
    if '瓶' in ctx:  return 'bottle'
    if prefix in ('CC','SH'): return 'component'
    return 'other'

def classify_file(rel):
    b = os.path.basename(rel)
    if rel.count(os.sep) == 0 and ('Invoice' in b or 'Group' in b): return 'invoice'
    if '玻璃瓶统一报价' in b: return 'glass'
    if '系列统一报价' in b:   return 'master'
    if os.sep + '系列' + os.sep in rel: return 'series'
    if os.sep + '单品报价' + os.sep in rel: return 'single'
    return 'other'

def file_date(b):
    m = re.search(r'(20\d{2})[-.](\d{1,2})[-.](\d{1,2})', b)
    return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}" if m else None

def series_from(*texts):
    for t in texts:
        if not t: continue
        m = re.search(r'H\d{3}', str(t).upper())
        if m: return m.group()
    return None

# --------------------------------------------------------------------------
def load_rows_xlsx(path):
    """{sheet_name: rows(list[list])}"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for ws in wb.worksheets:
        rows = []
        empty = 0
        for row in ws.iter_rows(max_row=min(ws.max_row, 800), values_only=True):
            if all(c is None for c in row):
                empty += 1
                if empty > 25: break
                rows.append(list(row)); continue
            empty = 0
            rows.append(list(row))
        out[ws.title] = rows
    wb.close()
    return out

def load_rows_xls(path):
    wb = xlrd.open_workbook(path)
    out = {}
    for sh in wb.sheets():
        if sh.nrows == 0: continue
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(min(sh.nrows, 800))]
        out[sh.name] = rows
    return out

# --------------------------------------------------------------------------
def detect_header(rows):
    """返回 (header_row_index, header_cells) —— 选标准 token 命中最多的前 8 行之一"""
    best_i, best_score, best = None, 0, None
    toks = ['容量','工艺','统一','编号','型号','供应商','图片','价格','配件','配套','物料','报价','规格','部件']
    for i, row in enumerate(rows[:8]):
        cells = [norm(c) for c in row]
        score = sum(1 for c in cells if any(t in c for t in toks))
        nonempty = sum(1 for c in cells if c)
        if score >= 2 and score > best_score and nonempty >= 3:
            best_i, best_score, best = i, score, row
    return best_i, best

def map_columns(header):
    """把表头映射到逻辑列索引"""
    cols = {'code':None,'capacity':None,'component':None,'process':None,
            'unified':None,'list':None,'supplier':None,'supplier_model':None,
            'remark':None,'image':None,'name':None,'wide':[]}
    unified_hdr = list_hdr = None
    for idx, raw in enumerate(header):
        h = norm(raw)
        if not h: continue
        is_supplier_model = ('供应商' in h and ('型号' in h or '编号' in h or '规格' in h)) \
            or any(sp in h and ('型号' in h or '编号' in h or '规格' in h) for sp in SUPPLIERS)
        if '图片' in h and cols['image'] is None: cols['image']=idx; continue
        if is_supplier_model and cols['supplier_model'] is None: cols['supplier_model']=idx; continue
        if ('供应商' in h) and ('型号' not in h) and cols['supplier'] is None: cols['supplier']=idx; continue
        if '备注' in h and cols['remark'] is None: cols['remark']=idx; continue
        if '物料名称' in h and cols['name'] is None: cols['name']=idx; continue
        if ('产品编号' in h or '李记型号' in h or '产品型号' in h or '物料' in h
            or (('编号' in h or '型号' in h) and '供应商' not in h) or h in ('产品','产品编号')) \
           and cols['code'] is None:
            cols['code']=idx; continue
        if ('容量' in h or '尺寸' in h or ('规格' in h and '型号' not in h)) and cols['capacity'] is None:
            cols['capacity']=idx; continue
        if ('配件' in h or '配套' in h or '结构' in h or '部件' in h or '编码名称' in h) and cols['component'] is None:
            cols['component']=idx; continue
        if '工艺' in h and cols['process'] is None: cols['process']=idx; continue
        if '门市' in h and cols['list'] is None: cols['list']=idx; list_hdr=raw; continue
        if '统一' in h and ('价' in h or '报价' in h) and cols['unified'] is None:
            cols['unified']=idx; unified_hdr=raw; continue
        # 其余：若含工艺关键词，作为 wide 价格列候选
        if any(k in h for k in PROC_KW):
            cols['wide'].append((idx, raw))
    cols['_unified_hdr']=unified_hdr; cols['_list_hdr']=list_hdr
    return cols

def col_has_price(rows, hidx, cidx):
    for r in rows[hidx+1:hidx+40]:
        if cidx < len(r):
            v = num(r[cidx])
            if v is not None and 0 < v < 200: return True
    return False

# --------------------------------------------------------------------------
def parse_table(rows, file_default_series=None, price_year_default=None):
    """通用解析 -> (items, codemap)  codemap: list[(row_index_1based, code, series)]"""
    hi, header = detect_header(rows)
    if hi is None: return [], []
    cols = map_columns(header)
    # 确认 wide 列确实有价格
    wide = [(i, raw) for (i, raw) in cols['wide'] if col_has_price(rows, hi, i)]
    has_long = cols['process'] is not None and (cols['unified'] is not None or cols['list'] is not None)
    if not has_long and not wide:
        return [], []
    year = find_year(cols.get('_unified_hdr'), cols.get('_list_hdr')) or price_year_default
    items, codemap = [], []
    cur = {'code':None,'cap':None,'series':file_default_series,'comp':None,'sup':None,'smodel':None}
    def g(row, idx):
        return row[idx] if (idx is not None and idx < len(row) and row[idx] not in (None,'')) else None
    for ri, row in enumerate(rows[hi+1:], start=hi+2):   # 1-based row number
        if all(c in (None,'') for c in row): continue
        code = g(row, cols['code'])
        cap  = g(row, cols['capacity'])
        comp = g(row, cols['component'])
        sup  = g(row, cols['supplier'])
        smod = g(row, cols['supplier_model'])
        rmk  = g(row, cols['remark'])
        # fill-down（仅在单元格像“真实编码”时更新 code，避免把备注/说明当成编码）
        if code is not None and is_codelike(code):
            cur['code']=re.sub(r'\s+',' ',str(code)).strip()
            s = series_from(code)
            if s: cur['series']=s
        if cap  is not None: cur['cap']=str(cap).strip()
        if comp is not None: cur['comp']=str(comp).strip()
        if sup  is not None and not is_num_only(sup):
            cur['sup']=re.sub(r'\s+',' ',str(sup)).strip()
        if smod is not None: cur['smodel']=str(smod).strip()
        # 记录该行的活动编码(供图片映射)
        if cur['code']:
            codemap.append((ri, cur['code'], cur['series']))
        base = dict(series=cur['series'], code=cur['code'], cap=cur['cap'], comp=cur['comp'],
                    sup=cur['sup'], smodel=cur['smodel'], remark=rmk, year=year, row=ri,
                    raw=rawjson(row))
        # LONG：本行有工艺 + 价格
        if has_long:
            proc = g(row, cols['process'])
            up = num(g(row, cols['unified'])) if cols['unified'] is not None else None
            lp = num(g(row, cols['list'])) if cols['list'] is not None else None
            if proc is not None and (up is not None or lp is not None):
                it = dict(base); it.update(process=str(proc).strip(), unified=up, list=lp, orient='long')
                items.append(it)
        # WIDE：每个工艺列一价
        for (i, raw) in wide:
            v = num(row[i]) if i < len(row) else None
            if v is not None and 0 < v < 200:
                it = dict(base); it.update(process=str(raw).replace('\n','').strip(),
                                           unified=v, list=None, orient='wide')
                items.append(it)
    return items, codemap

# --------------------------------------------------------------------------
def parse_glass(rows_by_sheet, fid, cur):
    for sheet, rows in rows_by_sheet.items():
        s = norm(sheet)
        if '库存' in s:
            for row in rows[2:]:
                if not row or row[0] in (None,''): continue
                cur.execute("INSERT INTO inventory(source_file_id,material_code,base_code,stock_qty,description,bottle_weight,stock_days,note) VALUES(?,?,?,?,?,?,?,?)",
                    (fid, row[0], row[1] if len(row)>1 else None, num(row[2]) if len(row)>2 else None,
                     row[3] if len(row)>3 else None, num(row[4]) if len(row)>4 else None,
                     num(row[5]) if len(row)>5 else None, row[6] if len(row)>6 else None))
            continue
        if any(k in s for k in ['乳液瓶','膏霜瓶','外购瓶','新开发','促销','加厚']):
            # 定位表头
            hi, header = detect_header(rows)
            if hi is None: continue
            H = [norm(c) for c in header]
            def find(*keys):
                for j,h in enumerate(H):
                    if any(k in h for k in keys): return j
                return None
            c_year=find('开发年份','年份'); c_code=find('物料','编码','编号'); c_name=find('名称')
            c_cap=find('容量'); c_wt=find('重量'); c_uni=find('统一'); c_bare=find('光瓶'); c_frost=find('蒙砂')
            for row in rows[hi+1:]:
                if not row or (c_code is not None and (c_code>=len(row) or row[c_code] in (None,''))): continue
                def gv(j): return row[j] if (j is not None and j<len(row)) else None
                cur.execute("INSERT INTO glass_bottles(source_file_id,sheet_name,dev_year,material_code,material_name,capacity,weight_g,unified_price,bare_price,frosting_fee,raw_json) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                    (fid, sheet, gv(c_year), str(gv(c_code)).strip() if gv(c_code) else None, gv(c_name),
                     str(gv(c_cap)) if gv(c_cap) is not None else None, num(gv(c_wt)),
                     num(gv(c_uni)), num(gv(c_bare)), num(gv(c_frost)),
                     rawjson(row)))

# --------------------------------------------------------------------------
def insert_items(cur, fid, sheet, items):
    for it in items:
        cur.execute("""INSERT INTO price_items
            (source_file_id,sheet_name,row_index,series_code,product_code,capacity,component,
             process_desc,unified_price,list_price,supplier,supplier_model,price_year,remark,orientation,raw_json)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (fid, sheet, it['row'], it['series'], it['code'], it['cap'], it['comp'],
             it['process'], it['unified'], it['list'], it['sup'], it['smodel'], it['year'],
             it['remark'], it['orient'], it['raw']))
        pid = cur.lastrowid
        for c in split_codes(it['code']):
            cur.execute("INSERT INTO price_item_code(price_item_id,code) VALUES(?,?)", (pid, c))

def map_image_code(codemap, anchor_row_1based):
    """取锚定行(含以上)最近的活动编码"""
    best = None
    for (ri, code, series) in codemap:
        if ri <= anchor_row_1based + 1:   # 允许图片略低于产品首行
            best = (code, series)
        else:
            break
    if best is None and codemap:
        best = (codemap[0][1], codemap[0][2])
    return best if best else (None, None)

def process_images(cur, fid, path, codemaps):
    imgs = imgmod.extract(path)
    n = 0
    for im in imgs:
        cm = codemaps.get(im['sheet_name'], [])
        arow = (im['from_row'] or 0) + 1
        code, series = map_image_code(cm, arow)
        safe = re.sub(r'[^\w\-]', '_', (code or 'UNMAPPED'))
        sub = '' if code else '_unmapped/'
        fname = f"{safe}__f{fid}_{im['media_name']}"
        os.makedirs(os.path.join(IMGDIR, sub), exist_ok=True)
        rel = f"images/{sub}{fname}"
        with open(os.path.join(DBDIR, rel), 'wb') as fh:
            fh.write(im['data'])
        cur.execute("""INSERT INTO images(source_file_id,sheet_name,anchor_row,anchor_col,product_code,series_code,media_name,ext,saved_path,mapped)
                       VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (fid, im['sheet_name'], arow, (im['from_col'] or 0)+1, code, series,
                     im['media_name'], im['ext'], rel, 1 if code else 0))
        n += 1
    return n

# --------------------------------------------------------------------------
def parse_invoices(cur):
    # 1) 根目录 IEB...xlsx (proforma invoice, EUR)
    p = os.path.join(ROOT, 'IEB252305LAPGroup-001.xlsx')
    if os.path.exists(p):
        fid = cur.execute("SELECT id FROM source_files WHERE rel_path=?", ('IEB252305LAPGroup-001.xlsx',)).fetchone()[0]
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True); ws = wb['CI']
        rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]; wb.close()
        cur.execute("""INSERT INTO invoices(source_file_id,doc_type,buyer,incoterm,currency,payment_term,total_amount,validity)
                       VALUES(?,?,?,?,?,?,?,?)""",
                    (fid,'proforma_invoice','Lap Group (Paris)','FOB Shanghai','EUR',
                     '20% down 80% against B/L', 117000, None))
        iid = cur.lastrowid
        for r in rows:
            txt = str(r[0]) if r and r[0] else ''
            if re.match(r'^\s*\d+\)', txt):
                qty=up=amt=None
                nums=[num(c) for c in r if num(c) is not None]
                # 结构: 描述..., qty, unit, total
                cur.execute("INSERT INTO invoice_lines(invoice_id,line_no,description,product_codes,quantity,unit_price,amount,raw_json) VALUES(?,?,?,?,?,?,?,?)",
                    (iid, txt.split(')')[0].strip(), txt.strip(), ','.join(split_codes(txt)),
                     nums[0] if len(nums)>=3 else None, nums[-2] if len(nums)>=2 else None,
                     nums[-1] if nums else None,
                     rawjson(r)))
    # 2) 根目录 Proforma Invoice.xls (quotation, USD)  —— 保留为报价单
    p2 = os.path.join(ROOT, 'Proforma Invoice.xls')
    if os.path.exists(p2):
        fid = cur.execute("SELECT id FROM source_files WHERE rel_path=?", ('Proforma Invoice.xls',)).fetchone()[0]
        wb = xlrd.open_workbook(p2); sh = wb.sheet_by_index(0)
        cur.execute("""INSERT INTO invoices(source_file_id,doc_type,buyer,incoterm,currency,payment_term,total_amount,validity)
                       VALUES(?,?,?,?,?,?,?,?)""",
                    (fid,'quotation',None,'FOB Guangzhou','USD','30% down, 70% T/T against BL',None,'2026-06-04'))
        iid = cur.lastrowid
        for r in range(6, sh.nrows):
            v0 = sh.cell_value(r,0)
            if isinstance(v0,(int,float)) and v0>0:
                desc = str(sh.cell_value(r,1))
                cur.execute("INSERT INTO invoice_lines(invoice_id,line_no,description,product_codes,quantity,unit_price,amount,raw_json) VALUES(?,?,?,?,?,?,?,?)",
                    (iid, str(v0), desc, ','.join(split_codes(desc)),
                     num(sh.cell_value(r,3)), num(sh.cell_value(r,4)), None,
                     rawjson([sh.cell_value(r,c) for c in range(sh.ncols)])))

# --------------------------------------------------------------------------
def seed_rules(cur):
    cur.executemany("INSERT INTO customer_markups(tier,markup_pct,note,confirmed) VALUES(?,?,?,?)",
        [('A',20,'通话中提到的示例加成，待核对',0),('B',30,'待核对',0),('C',50,'待核对',0)])
    cur.executemany("INSERT INTO qty_discount_tiers(min_qty,note,confirmed) VALUES(?,?,?)",
        [(5000,'基准价起点',0),(20000,'2万+ 有阶梯下降，比率待核对',0),(50000,'5万+ 进一步下降，待核对',0)])
    cur.executemany("INSERT INTO surface_fees(process,fee,unit,note,confirmed) VALUES(?,?,?,?,?)",
        [('普通印刷',0.20,'每次','通话口径，待核对',0),('烫金',0.25,'每次','通话口径，待核对',0),
         ('蒙砂/门杀',0.40,'每个','各表 0.15~0.4 不等，见 glass_bottles',0)])

def derive(cur):
    # products
    cur.execute("""INSERT OR IGNORE INTO products(code,prefix,product_type,capacity,default_supplier,n_price_rows)
        SELECT c.code,
               CASE WHEN c.code GLOB '[A-Z][0-9]*' THEN substr(c.code,1,1)
                    WHEN c.code GLOB '[A-Z][A-Z][0-9]*' THEN substr(c.code,1,2)
                    ELSE substr(c.code,1,2) END,
               NULL, NULL, NULL, COUNT(*)
        FROM price_item_code c GROUP BY c.code""")
    # 补 prefix/type/容量/供应商
    for row in cur.execute("SELECT code FROM products").fetchall():
        code = row[0]
        m = re.match(r'([A-Z]{1,3})', code); pref = m.group(1) if m else ''
        info = cur.execute("""SELECT pi.capacity, pi.supplier, pi.component, pi.process_desc
            FROM price_item_code pic JOIN price_items pi ON pi.id=pic.price_item_id
            WHERE pic.code=? LIMIT 1""",(code,)).fetchone()
        cap, sup, comp, proc = info if info else (None,None,None,None)
        cur.execute("UPDATE products SET prefix=?, product_type=?, capacity=?, default_supplier=? WHERE code=?",
                    (pref, infer_type(pref, code, '', comp or '', proc or ''), cap, sup, code))
    # glass 物料也进 products
    cur.execute("""INSERT OR IGNORE INTO products(code,prefix,product_type,name,capacity,weight_g,n_price_rows)
        SELECT material_code,
               CASE WHEN material_code GLOB '[A-Z][A-Z][0-9]*' THEN substr(material_code,1,2) ELSE substr(material_code,1,2) END,
               'bottle', material_name, capacity, weight_g, 0
        FROM glass_bottles WHERE material_code IS NOT NULL GROUP BY material_code""")
    # suppliers（归一到已知供应商名，去除型号/备注噪声）
    from collections import Counter
    cnt = Counter()
    for (s,) in cur.execute("SELECT supplier FROM price_items WHERE supplier IS NOT NULL AND supplier<>''"):
        hit = [sp for sp in SUPPLIERS if sp in s]
        for sp in (hit if hit else ['其他/未归类']):
            cnt[sp] += 1
    cur.executemany("INSERT INTO suppliers(name,n_items) VALUES(?,?)", list(cnt.items()))

# --------------------------------------------------------------------------
def main():
    if os.path.exists(DB): os.remove(DB)
    con = sqlite3.connect(DB); cur = con.cursor()
    with open(os.path.join(DBDIR,'schema.sql'),encoding='utf-8') as f:
        cur.executescript(f.read())

    files = []
    for dp,dn,fn in os.walk(ROOT):
        if 'pricing_db' in dp: continue
        for f in fn:
            if f.startswith('~$'): continue
            if f.lower().endswith(('.xlsx','.xls')):
                files.append(os.path.join(dp,f))
    files.sort()

    stats = dict(files=0, items=0, images=0, glass=0, inv=0)
    for path in files:
        rel = os.path.relpath(path, ROOT)
        cat = classify_file(rel)
        b = os.path.basename(rel)
        try:
            import zipfile
            imgcount = len([n for n in zipfile.ZipFile(path).namelist() if n.startswith('xl/media/')]) if path.endswith('.xlsx') else 0
        except Exception:
            imgcount = 0
        # register
        try:
            rows_by_sheet = load_rows_xlsx(path) if path.lower().endswith('.xlsx') else load_rows_xls(path)
        except Exception as e:
            print('  ! load fail', rel, e); rows_by_sheet = {}
        cur.execute("INSERT INTO source_files(rel_path,category,file_date,sheet_count,image_count) VALUES(?,?,?,?,?)",
                    (rel, cat, file_date(b), len(rows_by_sheet), imgcount))
        fid = cur.lastrowid
        stats['files'] += 1

        if cat == 'invoice':
            pass  # 单独处理
        elif cat == 'glass':
            before = cur.execute("SELECT COUNT(*) FROM glass_bottles").fetchone()[0]
            parse_glass(rows_by_sheet, fid, cur)
            stats['glass'] += cur.execute("SELECT COUNT(*) FROM glass_bottles").fetchone()[0]-before
        else:
            fdef_series = series_from(b)
            fyear = find_year(b)
            codemaps = {}
            for sheet, rows in rows_by_sheet.items():
                # 套系优先取“sheet 名”(主表每个 tab = 一个套系，如 H547统一报价)，
                # 其次才用文件名(单套系文件)。否则整份主表会被误标成文件名里的第一个 H###。
                sser = series_from(sheet) or fdef_series
                items, cmap = parse_table(rows, file_default_series=sser,
                                          price_year_default=fyear)
                insert_items(cur, fid, sheet, items)
                stats['items'] += len(items)
                if cmap: codemaps[sheet] = cmap
            # 图片
            if path.lower().endswith('.xlsx') and imgcount:
                stats['images'] += process_images(cur, fid, path, codemaps)

    parse_invoices(cur)
    seed_rules(cur)
    derive(cur)
    con.commit()

    # 汇总
    def q(s): return cur.execute(s).fetchone()[0]
    print("=== BUILD DONE ===")
    print("files           :", q("SELECT COUNT(*) FROM source_files"))
    print("price_items      :", q("SELECT COUNT(*) FROM price_items"))
    print("price_item_code  :", q("SELECT COUNT(*) FROM price_item_code"))
    print("products         :", q("SELECT COUNT(*) FROM products"))
    print("glass_bottles    :", q("SELECT COUNT(*) FROM glass_bottles"))
    print("inventory        :", q("SELECT COUNT(*) FROM inventory"))
    print("images           :", q("SELECT COUNT(*) FROM images"),
          " mapped:", q("SELECT COUNT(*) FROM images WHERE mapped=1"))
    print("invoices/lines   :", q("SELECT COUNT(*) FROM invoices"), "/", q("SELECT COUNT(*) FROM invoice_lines"))
    print("suppliers        :", q("SELECT COUNT(*) FROM suppliers"))
    con.close()

if __name__ == '__main__':
    main()
